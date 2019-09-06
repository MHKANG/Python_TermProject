#인구 데이터 분석 Module
from openpyxl import load_workbook, Workbook
import numpy as np
from matplotlib import pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import rc


def p_fixexcel(year):
    year = str(year)
    excelname = 'pinput_'+year+'.xlsx'
    gue = [] #구
    whole_countlist = [] # 계
    list_under = [] # 10세 미만
    list_10th = [] # 10대
    list_20th = [] # 20대
    list_30th = [] # 30대
    list_40th = [] # 40대
    list_50th = [] # 50대
    list_60th = [] # 60대
    list_70th = [] # 70대
    list_80th = [] # 80대
    list_90th = [] # 90대
    list_over = [] # 100세 이상


    hospitalEx = load_workbook(filename = excelname)
    sheet1 = hospitalEx['p'+year]
    
    wb = Workbook()
    sheetout = wb.active
    
    outfile = 'popmanu_'+year+'.xlsx'
    sheetout.title = year+'_pout'

    p_gue = p_cal_gue(sheet1, gue)
    p_whole = p_whole_count(sheet1, whole_countlist)
    p_under = p_under_count(sheet1, list_under)
    p_10th = p_10th_count(sheet1, list_10th)
    p_20th = p_20th_count(sheet1, list_20th)
    p_30th = p_30th_count(sheet1, list_30th)
    p_40th = p_40th_count(sheet1, list_40th)
    p_50th = p_50th_count(sheet1, list_50th)
    p_60th = p_60th_count(sheet1, list_60th)
    p_70th = p_70th_count(sheet1, list_70th)
    p_80th = p_80th_count(sheet1, list_80th)
    p_90th = p_90th_count(sheet1, list_90th)
    p_over = p_over_count(sheet1, list_over)

    
    sheetout.cell(1, column =1).value = "자치구"
    sheetout.cell(1, column =2).value = "총 인원 수"
    sheetout.cell(1, column =3).value = "10세 미만"
    sheetout.cell(1, column =4).value = "10대"
    sheetout.cell(1, column =5).value = "20대"
    sheetout.cell(1, column =6).value = "30대"
    sheetout.cell(1, column =7).value = "40대"
    sheetout.cell(1, column =8).value = "50대"
    sheetout.cell(1, column =9).value = "60대"
    sheetout.cell(1, column =10).value = "70대"
    sheetout.cell(1, column =11).value = "80대"
    sheetout.cell(1, column =12).value = "90대"
    sheetout.cell(1, column =13).value = "100세이상"

    for row_index in range(2, 28):
        sheetout.cell(row = row_index, column = 1).value = p_gue[row_index-2]
        sheetout.cell(row = row_index, column = 2).value = p_whole[row_index-2] 
        sheetout.cell(row = row_index, column = 3).value = p_under[row_index-2]
        sheetout.cell(row = row_index, column = 4).value = p_10th[row_index-2]
        sheetout.cell(row = row_index, column = 5).value = p_20th[row_index-2]
        sheetout.cell(row = row_index, column = 6).value = p_30th[row_index-2]
        sheetout.cell(row = row_index, column = 7).value = p_40th[row_index-2]
        sheetout.cell(row = row_index, column = 8).value = p_50th[row_index-2]
        sheetout.cell(row = row_index, column = 9).value = p_60th[row_index-2]
        sheetout.cell(row = row_index, column = 10).value = p_70th[row_index-2]
        sheetout.cell(row = row_index, column = 11).value = p_80th[row_index-2] 
        sheetout.cell(row = row_index, column = 12).value = p_90th[row_index-2]
        sheetout.cell(row = row_index, column = 13).value = p_over[row_index-2] 


        wb.save(filename = outfile)


def p_cal_gue (sheet, gue):
    for i in range(0,26):
        gue.append(sheet['B{}'.format(2+3*i)].value)
    return gue

def p_whole_count(sheet, whole_countlist):
    for i in range(0,26):
        whole_countlist.append(sheet['D{}'.format(2+3*i)].value)
    return whole_countlist

def p_under_count(sheet, list_under):
    list_u = ['E','F','G','H','I','J','K', 'L', 'M','N']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_under.append(result)
    return list_under

def p_10th_count(sheet, list_10th):
    list_u = ['O','P','Q','R','S','T','U', 'V', 'W','X']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_10th.append(result)
    return list_10th

def p_20th_count(sheet, list_20th):
    list_u = ['Y','Z','AA','AB','AC','AD','AE', 'AF', 'AG','AH']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_20th.append(result)
    return list_20th

def p_30th_count(sheet, list_30th):
    list_u = ['AI','AJ','AK','AL','AM','AN','AO', 'AP', 'AQ','AR']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_30th.append(result)
    return list_30th

def p_40th_count(sheet, list_40th):
    list_u = ['AS','AT','AU','AV','AW','AX','AY', 'AZ', 'BA','BB']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_40th.append(result)
    return list_40th

def p_50th_count(sheet, list_50th):
    list_u = ['BC','BD','BE','BF','BG','BH','BI', 'BJ', 'BK','BL']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_50th.append(result)
    return list_50th

def p_60th_count(sheet, list_60th):

    list_u = ['BM','BN','BO','BP','BQ','BR','BS', 'BT', 'BU','BV']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_60th.append(result)
    return list_60th

def p_70th_count(sheet, list_70th):
    list_u = ['BW','BX','BY','BZ','CA','CB','CC', 'CD', 'CE','CF']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_70th.append(result)
    return list_70th

def p_80th_count(sheet, list_80th):
    list_u = ['CG','CH','CI','CJ','CK','CL','CM', 'CN', 'CO','CP']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_80th.append(result)
    return list_80th

def p_90th_count(sheet, list_90th):
    list_u = ['CQ','CR','CS','CT','CU','CV','CW', 'CX', 'CY','CZ']
    for i in range(0,26):
        result = 0
        for j in list_u:
            result = result + sheet[j+str(2+3*i)].value
        list_90th.append(result)
    return list_90th

def p_over_count(sheet, list_over):
    for i in range(0,26):
        list_over.append(sheet['DA'+str(2+3*i)].value)
    return list_over

def toatal_data(year):
    year = str(year)
    excelname = 'popmanu_'+year+'.xlsx'
    gue = [] #구
    list_total = [] # 총 인원 수 리스트
    list_under = [] # 10세 미만
    list_10th = [] # 10대
    list_20th = [] # 20대
    list_30th = [] # 30대
    list_40th = [] # 40대
    list_50th = [] # 50대
    list_60th = [] # 60대
    list_70th = [] # 70대
    list_80th = [] # 80대
    list_90th = [] # 90대
    list_over = [] # 100세 이상


    popEx = load_workbook(filename = excelname)
    sheet1 = popEx[year+'_pout']
    #구에 관한 데이터를 저장하는 for 문
    for i in range(3, 28):
        gue.append(sheet1['A{}'.format(i)].value)
    #총 인원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_total.append(sheet1['B{}'.format(i)].value)
    #10세 미만 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_under.append(sheet1['C{}'.format(i)].value)
    #10대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_10th.append(sheet1['D{}'.format(i)].value)
    #20대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_20th.append(sheet1['E{}'.format(i)].value)
    #30대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_30th.append(sheet1['F{}'.format(i)].value)
    #40대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_40th.append(sheet1['G{}'.format(i)].value)
    #50대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_50th.append(sheet1['H{}'.format(i)].value)
    #60대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_60th.append(sheet1['I{}'.format(i)].value)
    #70대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_70th.append(sheet1['J{}'.format(i)].value)
    #80대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_80th.append(sheet1['K{}'.format(i)].value)
    #90대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_90th.append(sheet1['L{}'.format(i)].value)
    #100세 이상 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_over.append(sheet1['M{}'.format(i)].value)
    
    t_max_index = list_total.index(max(list_total))
    tu_max_index = list_under.index(max(list_under))
    t10_max_index = list_10th.index(max(list_10th))
    t20_max_index = list_20th.index(max(list_20th))
    t30_max_index = list_30th.index(max(list_30th))
    t40_max_index = list_40th.index(max(list_40th))
    t50_max_index = list_50th.index(max(list_50th))
    t60_max_index = list_60th.index(max(list_60th))
    t70_max_index = list_70th.index(max(list_70th))
    t80_max_index = list_80th.index(max(list_80th))
    t90_max_index = list_90th.index(max(list_90th))
    to_max_index = list_over.index(max(list_over))

    print(year,"'s 서울시 구별 인구 수 평균:", sheet1['B2'].value/25)
    print(year,"'s 서울시 가장 많은 인구를 가진 구 :", gue[t_max_index],":",max(list_total),"명")
    print(year,"'s 가장 많은 10세 미만 인구를 가진 구 :", gue[tu_max_index], ":", max(list_under),"명")
    print(year,"'s 가장 많은 10대를 가진 구 :", gue[t10_max_index], ":", max(list_10th),"명")
    print(year,"'s 가장 많은 20대를 가진 구 :", gue[t20_max_index], ":", max(list_20th),"명")
    print(year,"'s 가장 많은 30대를 가진 구 :", gue[t30_max_index], ":", max(list_30th),"명")
    print(year,"'s 가장 많은 40대를 가진 구 :", gue[t40_max_index], ":", max(list_40th),"명")
    print(year,"'s 가장 많은 50대를 가진 구 :", gue[t50_max_index], ":", max(list_50th),"명")
    print(year,"'s 가장 많은 60대를 가진 구 :", gue[t60_max_index], ":", max(list_60th),"명")
    print(year,"'s 가장 많은 70대를 가진 구 :", gue[t70_max_index], ":", max(list_70th),"명")
    print(year,"'s 가장 많은 80대를 가진 구 :", gue[t80_max_index], ":", max(list_80th),"명")
    print(year,"'s 가장 많은 90대를 가진 구 :", gue[t90_max_index], ":", max(list_90th),"명")
    print(year,"'s 가장 많은 100세이상 인구를 가진 구 :", gue[to_max_index], ":", max(list_over),"명")


def max_pop(year):
    year = str(year)
    excelname = 'popmanu_'+year+'.xlsx'
    gue = [] #구
    list_total = [] # 총 인원 수 리스트
    list_60th = [] # 60대
    list_70th = [] # 70대
    list_80th = [] # 80대
    list_90th = [] # 90대
    list_over = [] # 100세 이상
    list_order = [] # 60 대 이상 고령화 인구

    popEx = load_workbook(filename = excelname)
    sheet1 = popEx[year+'_pout']

    #구에 관한 데이터를 저장하는 for 문
    for i in range(3, 28):
        gue.append(sheet1['A{}'.format(i)].value)
    #총 인원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_total.append(sheet1['B{}'.format(i)].value)
    
     #60대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_60th.append(sheet1['I{}'.format(i)].value)
    #70대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_70th.append(sheet1['J{}'.format(i)].value)
    #80대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_80th.append(sheet1['K{}'.format(i)].value)
    #90대 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_90th.append(sheet1['L{}'.format(i)].value)
    #100세 이상 인원 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_over.append(sheet1['M{}'.format(i)].value)
    #60대 이상 고령화 인구를 저장하는 for 문
    for i in range(0, 25):
        result = list_60th[i] + list_70th[i] + list_80th[i] + list_90th[i]+list_over[i]
        list_order.append(result)
    
    #총 인구중 가장 많은 인구를 가진 구의 인덱스를 표시하는 값
    t_max_index = list_total.index(max(list_total))
    #60대 이상 고령화 인구를 가장 많이 가진 구의 인덱스를 표시하는 값
    t_omax_index = list_order.index(max(list_order))

    result = (gue[t_max_index], max(list_total), gue[t_omax_index], max(list_order))

    return result

def show_graph(list_year):
    
    path_font = 'C:/Windows/Fonts/HANBaeKM.ttf'
    fontporp1 = fm.FontProperties(fname = path_font).get_name()
    rc('font', family = fontporp1)
    y1_value = ()
    x1_name = ()
    y2_value = ()
    x2_name = ()
    for i in list_year:
        
        y1_value += (max_pop(i)[1],)
        y2_value += (max_pop(i)[3],)
        x1_name += (str(i) +"\n"+ max_pop(i)[0],)
        x2_name += (str(i) +"\n"+ max_pop(i)[2],)
        
    n1_groups = len(x1_name)
    n2_groups = len(x2_name)
    index1 = np.arange(n1_groups)
    index2 = np.arange(n2_groups)
    
    plt.subplot(1,2,1)
   
    plt.bar(index1, y1_value, tick_label= x1_name, align ='center',color ='c')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('인구 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 인구 구를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n1_groups)
    
    plt.subplot(1,2,2)
    
    plt.bar(index2, y2_value, tick_label= x2_name, align ='center',color='m')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('인구 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 고령화 인구를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n2_groups)
    plt.tight_layout()
    plt.show()