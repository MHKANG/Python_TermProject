#병원 데이터 분석 Module
from openpyxl import load_workbook, Workbook
import numpy as np
from matplotlib import pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import rc


def fixexcel(year):
    year = str(year)
    excelname = 'input_'+year+'.xlsx'
    gue = [] #구
    whole_countlist = [] # 계
    mut_hos = [] # 종합 병원
    hos = [] # 병원
    dent_hos = [] # 치과

    hospitalEx = load_workbook(filename = excelname)
    sheet1 = hospitalEx[year]
    
    wb = Workbook()
    sheetout = wb.active
    
    outfile = 'manu_'+year+'.xlsx'
    sheetout.title = year+'_out'
    
    c_gue = cal_gue(sheet1, gue)
    c_hos = cal_hos(sheet1, hos)
    c_mut = cal_mut_hos(sheet1, mut_hos)
    c_whole = cal_whole_count(sheet1, whole_countlist)
    if(year == '2017'):
        c_dent = cal_dent_hos2017(sheet1, dent_hos)
    else:
        c_dent = cal_dent_hos(sheet1, dent_hos)

    sheetout.cell(1, column =1).value = "자치구"
    sheetout.cell(1, column =2).value = "총 병원 수"
    sheetout.cell(1, column =3).value = "총 병상 수"
    sheetout.cell(1, column =4).value = "종합병원 수"
    sheetout.cell(1, column =5).value = "종합병원 병상 수"
    sheetout.cell(1, column =6).value = "병원 수"
    sheetout.cell(1, column =7).value = "병원 병상 수"
    sheetout.cell(1, column =8).value = "치과 수"
    sheetout.cell(1, column =9).value = "치과 병상 수"

    for row_index in range(2, 28):
        sheetout.cell(row = row_index, column = 1).value = c_gue[row_index-2]
        sheetout.cell(row = row_index, column = 2).value = c_whole[row_index-2][0]
        sheetout.cell(row = row_index, column = 3).value = c_whole[row_index-2][1]
        sheetout.cell(row = row_index, column = 4).value = c_mut[row_index-2][0]
        sheetout.cell(row = row_index, column = 5).value = c_mut[row_index-2][1]
        sheetout.cell(row = row_index, column = 6).value = c_hos[row_index-2][0]
        sheetout.cell(row = row_index, column = 7).value = c_hos[row_index-2][1]
        sheetout.cell(row = row_index, column = 8).value = c_dent[row_index-2][0]
        sheetout.cell(row = row_index, column = 9).value = c_dent[row_index-2][1]  
        wb.save(filename = outfile)


#종합 병원 추출 반복문
def cal_mut_hos(sheet1, mut_hos):
    for i in range(3,29):
        temp = sheet1['E{}'.format(i)].value
        temp2 = sheet1['F{}'.format(i)].value
        temp = str(temp)
        temp2 = str(temp2)
        if temp == '-' :
            temp = 0
        if temp2 == '-':
            temp2 = 0
        temp = int(temp)
        temp2 = int(temp2)
        mut_hos.append([temp,temp2])
    return mut_hos
#일반 병원
def cal_hos(sheet1, hos):
    for i in range(3,29):
        temp = sheet1['G{}'.format(i)].value
        temp2 = sheet1['H{}'.format(i)].value
        temp = str(temp)
        temp2 = str(temp2)
        if temp == '-' :
            temp = 0
        if temp2 == '-':
            temp2 = 0
        temp = int(temp)
        temp2 = int(temp2)
        hos.append([temp,temp2])
    return hos
#치과 2017버전 2017과 그 이전 데이터 값들의 set 형식이 일부 다른 점이 있어서 함수를 새로 만듬
def cal_dent_hos2017(sheet1, dent_hos):
    for i in range(3,29):
        temp = sheet1['V{}'.format(i)].value
        temp2 = sheet1['W{}'.format(i)].value
        temp = str(temp)
        temp2 = str(temp2)
        if temp == '-' :
            temp = 0
        if temp2 == '-':
            temp2 = 0
        temp = int(temp)
        temp2 = int(temp2)
        dent_hos.append([temp,temp2])
    return dent_hos

#치과 2017이 아닌 버전
def cal_dent_hos(sheet1, dent_hos):
    for i in range(3,29):
        temp = sheet1['W{}'.format(i)].value
        temp2 = sheet1['X{}'.format(i)].value
        temp = str(temp)
        temp2 = str(temp2)
        if temp == '-' :
            temp = 0
        if temp2 == '-':
            temp2 = 0
        temp = int(temp)
        temp2 = int(temp2)
        dent_hos.append([temp,temp2])
    return dent_hos
# 구 반복문
def cal_gue (sheet1, gue):
    for i in range(3,29):
        gue.append(sheet1['B{}'.format(i)].value)
    return gue

#전체 계 넣는 반복문
def cal_whole_count(sheet1, whole_countlist):
    for i in range (3,29):
        whole_countlist.append([sheet1['C{}'.format(i)].value,sheet1['D{}'.format(i)].value])
    return whole_countlist

def show_data(year):
    year = str(year)
    excelname = 'manu_'+year+'.xlsx'

    hospitalEx = load_workbook(filename = excelname)
    sheet1 = hospitalEx[year+'_out']

    result = 0
    mresult = 0
    hresult = 0
    dresult = 0
    
    t_gue = [] # 구 정보
    t_hos = [] # 총 병원
    t_mhos = [] # 종합 병원
    hos = [] #일반 병원
    t_dhos = [] # 치과
    #구 데이터 리스트 생성
    for i in range(3,28): 
        t_gue.append(sheet1['A{}'.format(i)].value)

    #총 병원 수 데이터 리스트 생성
    for i in range(3,28):
        result = result + sheet1['B{}'.format(i)].value
        t_hos.append(sheet1['B{}'.format(i)].value)
    
    #종합 병원 수 데이터 리스트 생성
    for i in range(3,28):
        mresult = mresult + sheet1['D{}'.format(i)].value
        t_mhos.append(sheet1['D{}'.format(i)].value)
    
    #일반 병원 수 데이터 리스트 생성
    for i in range(3,28):
        hresult = hresult + sheet1['F{}'.format(i)].value
        hos.append(sheet1['F{}'.format(i)].value)
    
    #치과 병원 수 데이터 리스트 생성
    for i in range(3,28):
        dresult = dresult + sheet1['H{}'.format(i)].value
        t_dhos.append(sheet1['H{}'.format(i)].value)

    t_max_index = t_hos.index(max(t_hos)) # 총 병원 수를 가진 구의 위치를 나타내는 index
    mh_max_index = t_mhos.index(max(t_mhos)) # 종합 병원
    h_max_index = hos.index(max(hos)) # 일반 병원
    dh_max_index = t_dhos.index(max(t_dhos)) # 치과 병원

    print(year, "'s 서울 시 병원 데이터 ")
    print("서울 시 병원 수 평균 : ")
    print("총 병원 수 평균 :", result/25)
    print("종합 병원 수 평균 :", mresult/25)
    print("일반 병원 수 평균 :", hresult/25)
    print("치과 병원 수 평균 :", dresult/25)

    print(year, "'s 서울 시 종류 별 가장 많은 병원을 가진 구 와 병원 수")
    print("총 병원 중 최대 수를 가진 구 와 수   : ", t_gue[t_max_index],"\t ",max(t_hos))
    print("종합 병원 중 최대 수를 가진 구 와 수 : ", t_gue[mh_max_index],"\t",max(t_mhos))
    print("일반 병원 중 최대 수를 가진 구 와 수 : ", t_gue[h_max_index],"\t ",max(hos) )
    print("치과 병원 중 최대 수를 가진 구 와 수 : ", t_gue[dh_max_index],"\t",max(t_dhos))
def max_hos(year):
    year = str(year)
    excelname = 'manu_'+year+'.xlsx'
    gue = [] #구
    list_total = [] # 총 병원 수 리스트
    list_mhos = [] # 종합 병원
    list_hos = [] # 일반 병원
    list_den = [] # 치과

    popEx = load_workbook(filename = excelname)
    sheet1 = popEx[year+'_out']
    #구에 관한 데이터를 저장하는 for 문
    for i in range(3, 28):
        gue.append(sheet1['A{}'.format(i)].value)
    
    #총 병원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_total.append(sheet1['B{}'.format(i)].value)
    
    #종합 병원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_mhos.append(sheet1['D{}'.format(i)].value)

    #일반 병원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_hos.append(sheet1['F{}'.format(i)].value)
    
    #치과 병원 수 데이터를 저장하는 for 문
    for i in range(3, 28):
        list_den.append(sheet1['H{}'.format(i)].value)
    
    #총 병원을 가장 많은 인구를 가진 구의 인덱스를 표시하는 값
    t_max_index = list_total.index(max(list_total))
    #종합 병원 가장 많이 가진 구의 인덱스를 표시 하는 값
    t_mhmax_index = list_mhos.index(max(list_mhos))
    #일반 병원 가장 많이 가진 구의 인덱스를 표시 하는 값
    t_hmax_index = list_hos.index(max(list_hos))
    #치과 병원 가장 많이 가진 구의 인덱스를 표시 하는 값
    t_dmax_index = list_den.index(max(list_den))
    
    result2 = (gue[t_max_index], max(list_total), gue[t_mhmax_index], max(list_mhos),gue[t_hmax_index], max(list_hos),gue[t_dmax_index],max(list_den))
    return result2

def show_graph(list_year):
    #폰트 위치 경로 만약 그래프의 글자가 깨진다면 원하는 폰트의 경로를 
    #path_font에 넣어서 적용시키면 됩니다.
    path_font = 'C:/Windows/Fonts/HANBaeKM.ttf'
    fontporp1 = fm.FontProperties(fname = path_font).get_name()
    rc('font', family = fontporp1)
    y1_value = ()
    x1_name = ()
    y2_value = ()
    x2_name = ()
    y3_value = ()
    x3_name = ()
    y4_value =()
    x4_name = ()
    for i in list_year:
        y1_value += (max_hos(i)[1],)
        y2_value += (max_hos(i)[3],)
        y3_value += (max_hos(i)[5],)
        y4_value += (max_hos(i)[7],)

        x1_name += (str(i) +"\n"+ max_hos(i)[0],)
        x2_name += (str(i) +"\n"+ max_hos(i)[2],)
        x3_name += (str(i) +"\n"+ max_hos(i)[4],)
        x4_name += (str(i) +"\n"+ max_hos(i)[6],)

        
    n1_groups = len(x1_name)
    n2_groups = len(x2_name)
    n3_groups = len(x3_name)
    n4_groups = len(x4_name)

    index1 = np.arange(n1_groups)
    index2 = np.arange(n2_groups)
    index3 = np.arange(n3_groups)
    index4 = np.arange(n4_groups)

    
    plt.subplot(2,2,1)
   
    plt.bar(index1, y1_value, tick_label= x1_name, align ='center',color ='c')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('병원 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 병원 수를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n1_groups)
    
    plt.subplot(2,2,2)
    
    plt.bar(index2, y2_value, tick_label= x2_name, align ='center',color='m')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('병원 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 종합 병원 수 를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n2_groups)

    plt.subplot(2,2,3)
    
    plt.bar(index3, y3_value, tick_label= x3_name, align ='center',color='b')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('병원 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 일반 병원 수 를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n3_groups)
    
    plt.subplot(2,2,4)
    
    plt.bar(index4, y4_value, tick_label= x4_name, align ='center',color='g')
    plt.xlabel('연도 : 구', fontproperties =fontporp1)
    plt.ylabel('병원 수', fontproperties =fontporp1)
    plt.title('연도 별 최대 치과 병원 수 를 가진 구 데이터', fontproperties =fontporp1)
    plt.xlim(-1, n4_groups)

    plt.tight_layout()
    plt.show()

    



